"""
Aman's Team — Reactivation SMS Agent
Flask app with scheduler + Twilio webhook + Dashboard + Broadcast + Add Lead
"""

from flask import Flask, request, jsonify, render_template_string, session, redirect
from twilio.twiml.messaging_response import MessagingResponse
from datetime import datetime
import pytz
import random
import time
import os
import threading
import pandas as pd

app = Flask(__name__)
app.secret_key = os.environ.get("DASHBOARD_SECRET", "aman-royal-lepage-2024")

conversations = {}
campaign_paused = False

EASTERN         = pytz.timezone("America/Toronto")
DELAY_MIN       = 45
DELAY_MAX       = 90
DAILY_LIMIT     = 50
SEND_HOUR_START = 9
SEND_HOUR_END   = 20
DASHBOARD_PASS  = os.environ.get("DASHBOARD_PASSWORD", "aman2024")


def is_sending_hours() -> bool:
    now = datetime.now(EASTERN)
    return SEND_HOUR_START <= now.hour < SEND_HOUR_END


def human_delay():
    delay = random.randint(DELAY_MIN, DELAY_MAX)
    print(f"[THROTTLE] Waiting {delay}s before next message...")
    time.sleep(delay)


def run_daily_campaign():
    global campaign_paused
    if campaign_paused:
        print("[SCHEDULER] Campaign is paused, skipping.")
        return
    if not is_sending_hours():
        print("[SCHEDULER] Outside sending hours, skipping.")
        return

    from lead_tracker import (load_leads, get_pending_leads, get_followup_leads, update_lead_sent)
    from message_templates import get_initial_message, get_followup_message
    from sms_sender import send_sms

    df = load_leads()
    sent_count = 0

    for _, lead in get_followup_leads(df).iterrows():
        if sent_count >= DAILY_LIMIT or campaign_paused:
            break
        phone   = str(lead["Phone (Formatted)"]).strip()
        phase   = str(lead.get("Phase", "Phase 1")).strip()
        message = get_followup_message(str(lead["First Name"]).strip(), phase)
        if send_sms(phone, message):
            update_lead_sent(phone, message)
            sent_count += 1
            if sent_count < DAILY_LIMIT:
                human_delay()

    for _, lead in get_pending_leads(df).iterrows():
        if sent_count >= DAILY_LIMIT or campaign_paused:
            break
        phone        = str(lead["Phone (Formatted)"]).strip()
        first_name   = str(lead["First Name"]).strip()
        buyer_seller = str(lead.get("Buyer/Seller", "")).strip()
        fav_city     = str(lead.get("Favorite City", "")).strip()
        phase        = str(lead.get("Phase", "Phase 1")).strip()
        message      = get_initial_message(first_name, buyer_seller, fav_city, phase)
        if send_sms(phone, message):
            update_lead_sent(phone, message)
            sent_count += 1
            if sent_count < DAILY_LIMIT:
                human_delay()

    print(f"[SCHEDULER] Done. Sent: {sent_count}")


def scheduler_loop():
    last_run_date = None
    while True:
        now = datetime.now(EASTERN)
        today = now.date()
        if now.hour == 10 and last_run_date != today:
            print(f"[SCHEDULER] Starting daily campaign at {now}")
            last_run_date = today
            try:
                run_daily_campaign()
            except Exception as e:
                print(f"[SCHEDULER ERROR] {e}")
        time.sleep(60 * 30)


def send_delayed_reply(to_number: str, ai_reply: str, delay_seconds: int):
    from sms_sender import send_sms
    print(f"[TYPING DELAY] Waiting {delay_seconds}s...")
    time.sleep(delay_seconds)
    send_sms(to_number, ai_reply)


def run_broadcast(message: str, phones: list):
    from sms_sender import send_sms
    sent = 0
    for phone in phones:
        if send_sms(phone, message):
            sent += 1
        delay = random.randint(DELAY_MIN, DELAY_MAX)
        time.sleep(delay)
    print(f"[BROADCAST] Done. Sent: {sent}/{len(phones)}")


# ── AUTH ──────────────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    error = ""
    if request.method == "POST":
        if request.form.get("password") == DASHBOARD_PASS:
            session["logged_in"] = True
            return redirect("/")
        error = "Wrong password. Try again."
    return render_template_string(LOGIN_HTML, error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")


def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect("/login")
        return f(*args, **kwargs)
    return decorated


# ── DASHBOARD ─────────────────────────────────────────────────────────────────

@app.route("/")
@login_required
def dashboard():
    return render_template_string(DASHBOARD_HTML)


# ── API ───────────────────────────────────────────────────────────────────────

@app.route("/api/stats")
@login_required
def api_stats():
    from lead_tracker import load_leads
    try:
        df = load_leads()
        return jsonify({
            "total":     len(df),
            "sent":      len(df[df["SMS Status"] == "Sent"]),
            "pending":   len(df[df["SMS Status"] == "Pending"]),
            "replied":   len(df[df["Reply Received"] == "Yes"]),
            "hot":       len(df[df["Lead Temperature"] == "Hot"]),
            "warm":      len(df[df["Lead Temperature"] == "Warm"]),
            "opted_out": len(df[df["SMS Status"] == "Opted Out"]),
            "paused":    campaign_paused
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/leads")
@login_required
def api_leads():
    from lead_tracker import load_leads
    try:
        df     = load_leads()
        search = request.args.get("search", "").lower()
        filt   = request.args.get("filter", "all")

        if search:
            mask = (
                df["First Name"].astype(str).str.lower().str.contains(search) |
                df["Last Name"].astype(str).str.lower().str.contains(search) |
                df["Phone (Formatted)"].astype(str).str.contains(search)
            )
            df = df[mask]

        if filt == "hot":       df = df[df["Lead Temperature"] == "Hot"]
        elif filt == "warm":    df = df[df["Lead Temperature"] == "Warm"]
        elif filt == "cold":    df = df[df["Lead Temperature"] == "Cold"]
        elif filt == "pending": df = df[df["SMS Status"] == "Pending"]
        elif filt == "replied": df = df[df["Reply Received"] == "Yes"]

        df    = df.fillna("")
        leads = df[["First Name","Last Name","Phone (Formatted)","Buyer/Seller",
                    "SMS Status","SMS Sent At","Reply Received","Reply Text",
                    "Lead Temperature","Favorite City","Pipeline Stage"]].to_dict(orient="records")
        return jsonify({"leads": leads, "total": len(leads)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/add_lead", methods=["POST"])
@login_required
def api_add_lead():
    from lead_tracker import load_leads, LEADS_FILE, SHEET_NAME
    from openpyxl import load_workbook
    data = request.json or {}

    def fmt(p):
        if not p: return ""
        n = str(p).strip().replace("-","").replace("(","").replace(")","").replace(" ","").replace("+","")
        if len(n)==10: return f"+1{n}"
        elif len(n)==11 and n.startswith("1"): return f"+{n}"
        return f"+{n}"

    first_name = str(data.get("first_name","")).strip()
    phone      = fmt(data.get("phone",""))

    if not first_name:
        return jsonify({"error": "First name is required"}), 400
    if not phone:
        return jsonify({"error": "Phone number is required"}), 400

    try:
        existing_df = load_leads()
        if phone in existing_df["Phone (Formatted)"].astype(str).str.strip().values:
            return jsonify({"error": "This phone number already exists in your leads"}), 400

        wb   = load_workbook(LEADS_FILE)
        ws   = wb[SHEET_NAME]
        hdrs = [cell.value for cell in ws[1]]

        row = {
            "First Name":         first_name,
            "Last Name":          str(data.get("last_name","")),
            "Phone (Formatted)":  phone,
            "Email":              str(data.get("email","")),
            "Buyer/Seller":       str(data.get("buyer_seller","Buyer")),
            "Favorite City":      str(data.get("city","")),
            "Notes":              str(data.get("notes","")),
            "Phase":              str(data.get("phase","Phase 1")),
            "Source":             "Manual",
            "SMS Status":         "Pending",
            "SMS Sent At":        "",
            "SMS Message Sent":   "",
            "Reply Received":     "No",
            "Reply Text":         "",
            "Lead Temperature":   "",
            "Follow Up Required": "",
            "Agent Notes":        ""
        }

        ws.append([row.get(h,"") for h in hdrs])
        wb.save(LEADS_FILE)
        return jsonify({"success": True, "message": f"{first_name} added successfully!"})

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/conversation/<path:phone>")
@login_required
def api_conversation(phone):
    from lead_tracker import load_leads
    history = conversations.get(phone, [])
    try:
        df  = load_leads()
        df["Phone (Formatted)"] = df["Phone (Formatted)"].astype(str).str.strip()
        row = df[df["Phone (Formatted)"] == phone.strip()]
        initial_msg = str(row.iloc[0].get("SMS Message Sent","")) if not row.empty else ""
        reply_text  = str(row.iloc[0].get("Reply Text",""))        if not row.empty else ""
        return jsonify({"history": history, "initial_message": initial_msg, "reply_text": reply_text})
    except:
        return jsonify({"history": history, "initial_message": "", "reply_text": ""})


@app.route("/api/reply", methods=["POST"])
@login_required
def api_manual_reply():
    from sms_sender import send_sms
    data    = request.json
    phone   = data.get("phone","")
    message = data.get("message","")
    if not phone or not message:
        return jsonify({"error": "Phone and message required"}), 400
    if phone not in conversations:
        conversations[phone] = []
    conversations[phone].append({"role":"assistant","content":message})
    return jsonify({"success": send_sms(phone, message)})


@app.route("/api/broadcast/preview", methods=["POST"])
@login_required
def api_broadcast_preview():
    from lead_tracker import load_leads
    try:
        df      = load_leads()
        filters = request.json or {}
        df      = apply_broadcast_filters(df, filters)
        return jsonify({"count": len(df), "sample": df["First Name"].head(5).tolist()})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/broadcast/send", methods=["POST"])
@login_required
def api_broadcast_send():
    from lead_tracker import load_leads
    from message_templates import get_broadcast_message
    try:
        data    = request.json or {}
        message = data.get("message","").strip()
        filters = data.get("filters", {})
        if not message:
            return jsonify({"error": "Message is required"}), 400

        df = load_leads()
        df = apply_broadcast_filters(df, filters)

        if df.empty:
            return jsonify({"error": "No leads match your filters"}), 400

        send_list = []
        for _, lead in df.iterrows():
            phone      = str(lead.get("Phone (Formatted)","")).strip()
            first_name = str(lead.get("First Name","")).strip()
            phase      = str(lead.get("Phase","Phase 1")).strip()
            if phone:
                personalized = get_broadcast_message(first_name, message, phase)
                send_list.append((phone, personalized))

        messages = {p: m for p, m in send_list}

        def send_all():
            from sms_sender import send_sms
            sent = 0
            for phone, msg in messages.items():
                if send_sms(phone, msg):
                    sent += 1
                time.sleep(random.randint(DELAY_MIN, DELAY_MAX))
            print(f"[BROADCAST] Complete. Sent: {sent}/{len(messages)}")

        t = threading.Thread(target=send_all)
        t.daemon = True
        t.start()

        return jsonify({"status": "sending", "total": len(send_list),
                        "message": f"Broadcasting to {len(send_list)} leads now!"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def apply_broadcast_filters(df, filters):
    if filters.get("phase"):
        phases = filters["phase"] if isinstance(filters["phase"], list) else [filters["phase"]]
        if "Phase" in df.columns:
            df = df[df["Phase"].isin(phases)]
    if filters.get("buyer_seller"):
        types = filters["buyer_seller"] if isinstance(filters["buyer_seller"], list) else [filters["buyer_seller"]]
        df = df[df["Buyer/Seller"].isin(types)]
    if filters.get("city") and filters["city"].strip():
        df = df[df["Favorite City"].astype(str).str.lower().str.contains(filters["city"].lower())]
    if filters.get("temperature"):
        temps = filters["temperature"] if isinstance(filters["temperature"], list) else [filters["temperature"]]
        df = df[df["Lead Temperature"].isin(temps)]
    if filters.get("status") == "pending":
        df = df[df["SMS Status"] == "Pending"]
    elif filters.get("status") == "contacted":
        df = df[df["SMS Status"] == "Sent"]
    return df


@app.route("/api/upload", methods=["POST"])
@login_required
def api_upload_leads():
    from lead_tracker import load_leads, LEADS_FILE, SHEET_NAME
    from openpyxl import load_workbook
    import io
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400
    file = request.files["file"]
    try:
        new_df = pd.read_excel(io.BytesIO(file.read()))
        col_map = {
            "first name":"First Name","last name":"Last Name",
            "phone":"Phone (Formatted)","cell phone":"Phone (Formatted)",
            "email":"Email","email address":"Email",
            "buyer/seller":"Buyer/Seller","city":"Favorite City","notes":"Notes"
        }
        new_df.columns = [col_map.get(c.lower().strip(), c) for c in new_df.columns]

        def fmt(p):
            if pd.isna(p): return ""
            n = str(p).strip().replace("-","").replace("(","").replace(")","").replace(" ","").replace("+","")
            if len(n)==10: return f"+1{n}"
            elif len(n)==11 and n.startswith("1"): return f"+{n}"
            return f"+{n}"

        if "Phone (Formatted)" in new_df.columns:
            new_df["Phone (Formatted)"] = new_df["Phone (Formatted)"].apply(fmt)

        existing_df     = load_leads()
        existing_phones = set(existing_df["Phone (Formatted)"].astype(str).str.strip())
        new_df          = new_df[~new_df["Phone (Formatted)"].isin(existing_phones)]

        if new_df.empty:
            return jsonify({"added":0,"message":"No new leads — all duplicates"})

        for col in ["SMS Status","SMS Sent At","SMS Message Sent","Reply Received",
                    "Reply Text","Lead Temperature","Follow Up Required","Agent Notes"]:
            new_df[col] = "" if col != "SMS Status" else "Pending"
        new_df["Reply Received"] = "No"

        wb   = load_workbook(LEADS_FILE)
        ws   = wb[SHEET_NAME]
        hdrs = [cell.value for cell in ws[1]]
        for _, row in new_df.iterrows():
            ws.append([("" if pd.isna(row.get(h,"")) else row.get(h,"")) for h in hdrs])
        wb.save(LEADS_FILE)
        return jsonify({"added":len(new_df),"message":f"{len(new_df)} new leads added!"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/pause", methods=["POST"])
@login_required
def api_pause():
    global campaign_paused
    campaign_paused = not campaign_paused
    return jsonify({"status":"paused" if campaign_paused else "resumed","paused":campaign_paused})


@app.route("/api/zapier", methods=["POST"])
def api_zapier():
    from lead_tracker import load_leads, LEADS_FILE, SHEET_NAME
    from openpyxl import load_workbook
    data = request.json or request.form.to_dict()

    def fmt(p):
        if not p: return ""
        n = str(p).strip().replace("-","").replace("(","").replace(")","").replace(" ","").replace("+","")
        if len(n)==10: return f"+1{n}"
        elif len(n)==11 and n.startswith("1"): return f"+{n}"
        return f"+{n}"

    phone      = fmt(data.get("phone") or data.get("cell_phone") or data.get("phone_number",""))
    first_name = str(data.get("first_name") or data.get("name","")).strip()
    if not phone or not first_name:
        return jsonify({"error":"first_name and phone required"}), 400

    existing_df = load_leads()
    if phone in existing_df["Phone (Formatted)"].astype(str).values:
        return jsonify({"status":"duplicate"})

    wb   = load_workbook(LEADS_FILE)
    ws   = wb[SHEET_NAME]
    hdrs = [cell.value for cell in ws[1]]
    row  = {"First Name":first_name,"Last Name":str(data.get("last_name","")),
            "Phone (Formatted)":phone,"Email":str(data.get("email","")),
            "Buyer/Seller":str(data.get("buyer_seller","Buyer")),
            "Favorite City":str(data.get("city","")),
            "Notes":str(data.get("notes") or data.get("ad_notes","")),
            "Phase":"Phase 1","Source":"Zapier",
            "SMS Status":"Pending","SMS Sent At":"","SMS Message Sent":"",
            "Reply Received":"No","Reply Text":"","Lead Temperature":"",
            "Follow Up Required":"","Agent Notes":""}
    ws.append([row.get(h,"") for h in hdrs])
    wb.save(LEADS_FILE)
    return jsonify({"status":"success","message":f"Lead {first_name} added"})


# ── WEBHOOK ───────────────────────────────────────────────────────────────────

@app.route("/webhook/sms", methods=["POST"])
def sms_webhook():
    from lead_tracker import update_lead_reply, update_lead_optout, load_leads
    from sms_sender import generate_ai_reply, classify_lead_temperature
    from message_templates import get_system_prompt

    incoming_msg = request.form.get("Body","").strip()
    from_number  = request.form.get("From","").strip()
    resp         = MessagingResponse()

    print(f"[INCOMING] {from_number}: {incoming_msg}")

    if incoming_msg.upper() in ["STOP","UNSUBSCRIBE","CANCEL","QUIT","END"]:
        update_lead_optout(from_number)
        resp.message("You've been unsubscribed. You won't receive any more messages from us. Take care!")
        return str(resp)

    if from_number not in conversations:
        conversations[from_number] = []

    try:
        df    = load_leads()
        df["Phone (Formatted)"] = df["Phone (Formatted)"].astype(str).str.strip()
        row   = df[df["Phone (Formatted)"] == from_number.strip()]
        phase = str(row.iloc[0].get("Phase","Phase 1")) if not row.empty else "Phase 1"
    except:
        phase = "Phase 1"

    temperature   = classify_lead_temperature(incoming_msg)
    update_lead_reply(from_number, incoming_msg, temperature)
    system_prompt = get_system_prompt(phase)
    ai_reply      = generate_ai_reply(conversations[from_number], incoming_msg, system_prompt)
    words         = len(ai_reply.split())
    typing_delay  = random.randint(20,45) + (words//5)

    t = threading.Thread(target=send_delayed_reply, args=(from_number, ai_reply, typing_delay))
    t.daemon = True
    t.start()
    return str(resp)


@app.route("/health")
def health():
    return jsonify({"status":"running","agent":"Aman Reactivation Bot"})


@app.route("/trigger")
@login_required
def manual_trigger():
    t = threading.Thread(target=run_daily_campaign)
    t.daemon = True
    t.start()
    return jsonify({"status":"campaign triggered"})


# ── HTML ──────────────────────────────────────────────────────────────────────

LOGIN_HTML = """<!DOCTYPE html><html><head><title>Aman's Agent</title>
<style>*{margin:0;padding:0;box-sizing:border-box}body{background:#0f172a;display:flex;align-items:center;justify-content:center;height:100vh;font-family:Arial,sans-serif}.card{background:#1e293b;padding:40px;border-radius:12px;width:360px;text-align:center}h2{color:#fff;margin-bottom:8px}p{color:#94a3b8;margin-bottom:24px;font-size:14px}input{width:100%;padding:12px;border-radius:8px;border:1px solid #334155;background:#0f172a;color:#fff;font-size:15px;margin-bottom:16px}button{width:100%;padding:12px;background:#3b82f6;color:#fff;border:none;border-radius:8px;font-size:15px;cursor:pointer;font-weight:bold}button:hover{background:#2563eb}.error{color:#f87171;font-size:13px;margin-bottom:12px}</style></head>
<body><div class="card"><h2>🏠 Aman's Reactivation Agent</h2><p>Royal LePage — Command Centre</p>
{% if error %}<div class="error">{{ error }}</div>{% endif %}
<form method="POST"><input type="password" name="password" placeholder="Enter password" autofocus><button type="submit">Login</button></form>
</div></body></html>"""

DASHBOARD_HTML = """<!DOCTYPE html><html><head><title>Aman's Agent</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{background:#0f172a;color:#e2e8f0;font-family:Arial,sans-serif}
.header{background:#1e293b;padding:16px 24px;display:flex;align-items:center;justify-content:space-between;border-bottom:1px solid #334155}
.header h1{font-size:18px;color:#fff}
.btn{padding:8px 16px;border-radius:8px;border:none;cursor:pointer;font-size:13px;font-weight:bold;transition:opacity .2s}
.btn:hover{opacity:.85}
.btn-blue{background:#3b82f6;color:#fff}
.btn-green{background:#22c55e;color:#fff}
.btn-yellow{background:#f59e0b;color:#fff}
.btn-gray{background:#475569;color:#fff}
.btn-purple{background:#8b5cf6;color:#fff}
.btn-teal{background:#0d9488;color:#fff}
.header-right{display:flex;gap:10px;align-items:center}
.stats{display:grid;grid-template-columns:repeat(6,1fr);gap:12px;padding:20px 24px}
.stat-card{background:#1e293b;border-radius:10px;padding:16px;text-align:center;border:1px solid #334155}
.stat-card .num{font-size:28px;font-weight:bold;margin-bottom:4px}
.stat-card .label{font-size:12px;color:#94a3b8}
.hot-num{color:#f87171}.warm-num{color:#fbbf24}.sent-num{color:#60a5fa}.replied-num{color:#34d399}
.tabs{display:flex;gap:0;padding:0 24px;border-bottom:1px solid #334155}
.tab{padding:12px 20px;cursor:pointer;font-size:13px;font-weight:bold;color:#64748b;border-bottom:2px solid transparent}
.tab.active{color:#3b82f6;border-bottom-color:#3b82f6}
.tab-content{display:none;padding:20px 24px}
.tab-content.active{display:block}
.main{display:grid;grid-template-columns:1fr 380px;gap:16px}
.panel{background:#1e293b;border-radius:10px;border:1px solid #334155}
.panel-header{padding:14px 16px;border-bottom:1px solid #334155;display:flex;align-items:center;justify-content:space-between}
.panel-header h3{font-size:14px;font-weight:bold}
.filters{display:flex;gap:8px;padding:12px 16px;border-bottom:1px solid #334155;flex-wrap:wrap}
.filter-btn{padding:5px 12px;border-radius:20px;border:1px solid #334155;background:transparent;color:#94a3b8;cursor:pointer;font-size:12px}
.filter-btn.active{background:#3b82f6;color:#fff;border-color:#3b82f6}
.search-box{padding:12px 16px;border-bottom:1px solid #334155}
.search-box input{width:100%;padding:8px 12px;border-radius:8px;border:1px solid #334155;background:#0f172a;color:#fff;font-size:13px}
.leads-table{overflow-y:auto;max-height:450px}
table{width:100%;border-collapse:collapse;font-size:13px}
th{padding:10px 12px;text-align:left;color:#94a3b8;font-size:11px;text-transform:uppercase;position:sticky;top:0;background:#1e293b;border-bottom:1px solid #334155}
td{padding:10px 12px;border-bottom:1px solid #1a2332;cursor:pointer}
tr:hover td{background:#263548}tr.selected td{background:#1d3a5e}
.badge{padding:3px 8px;border-radius:12px;font-size:11px;font-weight:bold}
.badge-hot{background:#fee2e2;color:#dc2626}.badge-warm{background:#fef9c3;color:#b45309}
.badge-cold{background:#e0e7ff;color:#4338ca}.badge-pending{background:#f1f5f9;color:#475569}
.badge-sent{background:#dbeafe;color:#1d4ed8}.badge-optout{background:#fee2e2;color:#9f1239}
.badge-p1{background:#dcfce7;color:#15803d}.badge-p2{background:#fef9c3;color:#92400e}.badge-p3{background:#fee2e2;color:#991b1b}
.conv-panel{display:flex;flex-direction:column}
.conv-messages{flex:1;overflow-y:auto;padding:16px;max-height:350px;min-height:150px}
.msg{margin-bottom:12px}.msg-out{text-align:right}
.msg-bubble{display:inline-block;padding:8px 12px;border-radius:12px;font-size:13px;max-width:85%;line-height:1.4}
.msg-out .msg-bubble{background:#3b82f6;color:#fff;border-radius:12px 12px 2px 12px}
.msg-in .msg-bubble{background:#334155;color:#e2e8f0;border-radius:12px 12px 12px 2px}
.msg-label{font-size:11px;color:#64748b;margin-bottom:3px}
.reply-box{padding:12px 16px;border-top:1px solid #334155}
.reply-box textarea{width:100%;padding:10px;border-radius:8px;border:1px solid #334155;background:#0f172a;color:#fff;font-size:13px;resize:none;height:70px;margin-bottom:8px}
.upload-area{padding:16px}
.upload-label{display:block;padding:16px;border:2px dashed #334155;border-radius:8px;text-align:center;cursor:pointer;color:#94a3b8;font-size:13px}
.upload-label:hover{border-color:#3b82f6;color:#3b82f6}
.broadcast-box{padding:20px}
.broadcast-box textarea{width:100%;padding:12px;border-radius:8px;border:1px solid #334155;background:#0f172a;color:#fff;font-size:13px;resize:none;height:100px;margin-bottom:16px}
.filter-group{margin-bottom:16px}
.filter-group label{display:block;font-size:12px;color:#94a3b8;margin-bottom:8px;font-weight:bold;text-transform:uppercase}
.checkbox-group{display:flex;flex-wrap:wrap;gap:8px}
.checkbox-item{display:flex;align-items:center;gap:6px;font-size:13px;cursor:pointer}
.checkbox-item input{cursor:pointer;accent-color:#3b82f6}
.city-input{width:100%;padding:8px 12px;border-radius:8px;border:1px solid #334155;background:#0f172a;color:#fff;font-size:13px;margin-bottom:16px}
.preview-box{background:#0f172a;border-radius:8px;padding:12px;margin-bottom:16px;font-size:13px;color:#94a3b8;border:1px solid #334155}
.preview-box strong{color:#22c55e}
.toast{position:fixed;bottom:24px;right:24px;background:#22c55e;color:#fff;padding:12px 20px;border-radius:8px;font-size:13px;font-weight:bold;display:none;z-index:999}
.toast.error{background:#ef4444}
.status-dot{width:8px;height:8px;border-radius:50%;display:inline-block;margin-right:6px}
.dot-green{background:#22c55e}.dot-red{background:#ef4444}.dot-yellow{background:#f59e0b}
.modal-overlay{position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,.7);z-index:1000;display:none;align-items:center;justify-content:center}
.modal-overlay.open{display:flex}
.modal{background:#1e293b;border-radius:12px;padding:28px;width:480px;border:1px solid #334155;max-height:90vh;overflow-y:auto}
.modal h3{font-size:16px;font-weight:bold;margin-bottom:20px;color:#fff}
.form-row{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:12px}
.form-group{margin-bottom:12px}
.form-group label{display:block;font-size:12px;color:#94a3b8;margin-bottom:6px;font-weight:bold;text-transform:uppercase}
.form-group input,.form-group select,.form-group textarea{width:100%;padding:10px 12px;border-radius:8px;border:1px solid #334155;background:#0f172a;color:#fff;font-size:13px}
.form-group textarea{resize:none;height:60px}
.modal-actions{display:flex;gap:10px;margin-top:20px}
.form-group select option{background:#1e293b}
</style></head>
<body>

<!-- ADD LEAD MODAL -->
<div class="modal-overlay" id="add-lead-modal">
  <div class="modal">
    <h3>➕ Add New Lead</h3>
    <div class="form-row">
      <div class="form-group">
        <label>First Name *</label>
        <input id="al-firstname" placeholder="John">
      </div>
      <div class="form-group">
        <label>Last Name</label>
        <input id="al-lastname" placeholder="Smith">
      </div>
    </div>
    <div class="form-row">
      <div class="form-group">
        <label>Phone Number *</label>
        <input id="al-phone" placeholder="647-123-4567">
      </div>
      <div class="form-group">
        <label>Email</label>
        <input id="al-email" placeholder="john@email.com">
      </div>
    </div>
    <div class="form-row">
      <div class="form-group">
        <label>Buyer / Seller</label>
        <select id="al-type">
          <option value="Buyer">Buyer</option>
          <option value="Seller">Seller</option>
          <option value="Both">Both</option>
          <option value="Neither">Neither</option>
        </select>
      </div>
      <div class="form-group">
        <label>Phase</label>
        <select id="al-phase">
          <option value="Phase 1">Phase 1 — Recent (0-2 yrs)</option>
          <option value="Phase 2">Phase 2 — Warm (2-5 yrs)</option>
          <option value="Phase 3">Phase 3 — Cold (5+ yrs)</option>
        </select>
      </div>
    </div>
    <div class="form-group">
      <label>City</label>
      <input id="al-city" placeholder="Brampton, Mississauga...">
    </div>
    <div class="form-group">
      <label>Notes</label>
      <textarea id="al-notes" placeholder="Any notes about this lead..."></textarea>
    </div>
    <div class="modal-actions">
      <button class="btn btn-gray" onclick="closeAddLead()" style="flex:1;">Cancel</button>
      <button class="btn btn-teal" onclick="submitAddLead()" style="flex:2;">➕ Add Lead</button>
    </div>
  </div>
</div>

<div class="header">
  <div><h1>🏠 Aman's Reactivation Agent</h1><span id="agent-status"><span class="status-dot dot-yellow"></span>Loading...</span></div>
  <div class="header-right">
    <button class="btn btn-teal" onclick="openAddLead()">➕ Add Lead</button>
    <button class="btn btn-green" onclick="triggerCampaign()">🚀 Launch Campaign</button>
    <button class="btn btn-yellow" id="pause-btn" onclick="togglePause()">⏸ Pause</button>
    <button class="btn btn-gray" onclick="location='/logout'">Logout</button>
  </div>
</div>

<div class="stats">
  <div class="stat-card"><div class="num" id="s-total">—</div><div class="label">Total Leads</div></div>
  <div class="stat-card"><div class="num sent-num" id="s-sent">—</div><div class="label">SMS Sent</div></div>
  <div class="stat-card"><div class="num" id="s-pending">—</div><div class="label">Pending</div></div>
  <div class="stat-card"><div class="num replied-num" id="s-replied">—</div><div class="label">Replied</div></div>
  <div class="stat-card"><div class="num hot-num" id="s-hot">—</div><div class="label">🔥 Hot Leads</div></div>
  <div class="stat-card"><div class="num warm-num" id="s-warm">—</div><div class="label">🌡️ Warm Leads</div></div>
</div>

<div class="tabs">
  <div class="tab active" onclick="switchTab('leads',this)">📋 All Leads</div>
  <div class="tab" onclick="switchTab('broadcast',this)">📢 Broadcast</div>
  <div class="tab" onclick="switchTab('upload',this)">📁 Upload Leads</div>
</div>

<!-- LEADS TAB -->
<div class="tab-content active" id="tab-leads">
  <div class="main">
    <div class="panel">
      <div class="panel-header"><h3>All Leads</h3><span id="leads-count" style="font-size:12px;color:#94a3b8;"></span></div>
      <div class="filters">
        <button class="filter-btn active" onclick="setFilter('all',this)">All</button>
        <button class="filter-btn" onclick="setFilter('hot',this)">🔥 Hot</button>
        <button class="filter-btn" onclick="setFilter('warm',this)">🌡️ Warm</button>
        <button class="filter-btn" onclick="setFilter('replied',this)">💬 Replied</button>
        <button class="filter-btn" onclick="setFilter('pending',this)">⏳ Pending</button>
      </div>
      <div class="search-box"><input id="search-input" placeholder="🔍 Search by name or phone..." oninput="searchLeads()"></div>
      <div class="leads-table"><table><thead><tr><th>Name</th><th>Phone</th><th>Type</th><th>Phase</th><th>Status</th><th>Temp</th></tr></thead>
      <tbody id="leads-tbody"><tr><td colspan="6" style="padding:30px;text-align:center;color:#475569;">Loading...</td></tr></tbody></table></div>
    </div>
    <div style="display:flex;flex-direction:column;gap:16px;">
      <div class="panel conv-panel" style="flex:1;">
        <div class="panel-header"><h3 id="conv-title">Select a lead</h3></div>
        <div class="conv-messages" id="conv-messages"><div style="padding:30px;text-align:center;color:#475569;font-size:13px;">Click any lead to view conversation</div></div>
        <div class="reply-box" id="reply-box" style="display:none;">
          <textarea id="reply-text" placeholder="Type your message as Sarah..."></textarea>
          <button class="btn btn-blue" onclick="sendManualReply()" style="width:100%;">Send as Sarah</button>
        </div>
      </div>
    </div>
  </div>
</div>

<!-- BROADCAST TAB -->
<div class="tab-content" id="tab-broadcast">
  <div style="max-width:800px;">
    <div class="panel">
      <div class="panel-header"><h3>📢 Broadcast Message</h3><span style="font-size:12px;color:#94a3b8;">Send a property or announcement to selected leads</span></div>
      <div class="broadcast-box">
        <div class="filter-group">
          <label>Your Message</label>
          <textarea id="broadcast-msg" placeholder="e.g. We just listed a stunning 4-bed detached in Brampton for $899k — perfect for families. Would you like to see it? Reply YES and I'll send you the details!"></textarea>
        </div>
        <div class="filter-group">
          <label>Phase</label>
          <div class="checkbox-group">
            <label class="checkbox-item"><input type="checkbox" class="bc-phase" value="Phase 1" checked> Phase 1 (0-2 yrs)</label>
            <label class="checkbox-item"><input type="checkbox" class="bc-phase" value="Phase 2"> Phase 2 (2-5 yrs)</label>
            <label class="checkbox-item"><input type="checkbox" class="bc-phase" value="Phase 3"> Phase 3 (5+ yrs)</label>
          </div>
        </div>
        <div class="filter-group">
          <label>Buyer / Seller Type</label>
          <div class="checkbox-group">
            <label class="checkbox-item"><input type="checkbox" class="bc-type" value="Buyer" checked> Buyers</label>
            <label class="checkbox-item"><input type="checkbox" class="bc-type" value="Seller" checked> Sellers</label>
            <label class="checkbox-item"><input type="checkbox" class="bc-type" value="Both" checked> Both</label>
            <label class="checkbox-item"><input type="checkbox" class="bc-type" value="Neither"> Neither</label>
          </div>
        </div>
        <div class="filter-group">
          <label>Lead Temperature</label>
          <div class="checkbox-group">
            <label class="checkbox-item"><input type="checkbox" class="bc-temp" value="Hot" checked> 🔥 Hot</label>
            <label class="checkbox-item"><input type="checkbox" class="bc-temp" value="Warm" checked> 🌡️ Warm</label>
            <label class="checkbox-item"><input type="checkbox" class="bc-temp" value="Cold" checked> ❄️ Cold</label>
            <label class="checkbox-item"><input type="checkbox" class="bc-temp" value=""> Not yet rated</label>
          </div>
        </div>
        <div class="filter-group">
          <label>Filter by City (optional)</label>
          <input class="city-input" id="bc-city" placeholder="e.g. Brampton, Mississauga...">
        </div>
        <div class="preview-box" id="broadcast-preview">Click "Preview" to see how many leads will receive this message.</div>
        <div style="display:flex;gap:10px;">
          <button class="btn btn-gray" onclick="previewBroadcast()" style="flex:1;">👁️ Preview Count</button>
          <button class="btn btn-purple" onclick="sendBroadcast()" style="flex:2;">📢 Send Broadcast</button>
        </div>
      </div>
    </div>
  </div>
</div>

<!-- UPLOAD TAB -->
<div class="tab-content" id="tab-upload">
  <div style="max-width:600px;">
    <div class="panel">
      <div class="panel-header"><h3>📁 Upload New Leads</h3></div>
      <div class="upload-area">
        <label class="upload-label" for="file-upload">
          📂 Click to upload Excel/CSV leads file<br>
          <small style="color:#64748b;">Duplicates removed automatically</small>
        </label>
        <input type="file" id="file-upload" style="display:none;" accept=".xlsx,.csv" onchange="uploadLeads(this)">
      </div>
    </div>
  </div>
</div>

<div class="toast" id="toast"></div>

<script>
let currentFilter='all',currentPhone=null,searchTimer=null;

function switchTab(tab,el){
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  document.querySelectorAll('.tab-content').forEach(t=>t.classList.remove('active'));
  el.classList.add('active');
  document.getElementById('tab-'+tab).classList.add('active');
}

function openAddLead(){document.getElementById('add-lead-modal').classList.add('open');}
function closeAddLead(){document.getElementById('add-lead-modal').classList.remove('open');clearAddForm();}
function clearAddForm(){
  ['al-firstname','al-lastname','al-phone','al-email','al-city','al-notes'].forEach(id=>document.getElementById(id).value='');
  document.getElementById('al-type').value='Buyer';
  document.getElementById('al-phase').value='Phase 1';
}

async function submitAddLead(){
  const first_name=document.getElementById('al-firstname').value.trim();
  const phone=document.getElementById('al-phone').value.trim();
  if(!first_name||!phone){showToast('First name and phone are required',true);return;}
  const d=await(await fetch('/api/add_lead',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({
    first_name,
    last_name:document.getElementById('al-lastname').value.trim(),
    phone,
    email:document.getElementById('al-email').value.trim(),
    buyer_seller:document.getElementById('al-type').value,
    phase:document.getElementById('al-phase').value,
    city:document.getElementById('al-city').value.trim(),
    notes:document.getElementById('al-notes').value.trim()
  })})).json();
  if(d.error){showToast(d.error,true);}
  else{showToast('✅ '+d.message);closeAddLead();loadStats();loadLeads();}
}

async function loadStats(){
  const d=await(await fetch('/api/stats')).json();
  document.getElementById('s-total').textContent=d.total||0;
  document.getElementById('s-sent').textContent=d.sent||0;
  document.getElementById('s-pending').textContent=d.pending||0;
  document.getElementById('s-replied').textContent=d.replied||0;
  document.getElementById('s-hot').textContent=d.hot||0;
  document.getElementById('s-warm').textContent=d.warm||0;
  const st=document.getElementById('agent-status'),pb=document.getElementById('pause-btn');
  if(d.paused){st.innerHTML='<span class="status-dot dot-red"></span>Paused';pb.textContent='▶ Resume';pb.className='btn btn-green';}
  else{st.innerHTML='<span class="status-dot dot-green"></span>Agent Running';pb.textContent='⏸ Pause';pb.className='btn btn-yellow';}
}

async function loadLeads(){
  const s=document.getElementById('search-input').value;
  const d=await(await fetch(`/api/leads?filter=${currentFilter}&search=${encodeURIComponent(s)}`)).json();
  document.getElementById('leads-count').textContent=`${d.total} leads`;
  const tb=document.getElementById('leads-tbody');
  if(!d.leads||!d.leads.length){tb.innerHTML='<tr><td colspan="6" style="padding:30px;text-align:center;color:#475569;">No leads found</td></tr>';return;}
  tb.innerHTML=d.leads.map(l=>{
    const name=`${l['First Name']} ${l['Last Name']}`.trim(),phone=l['Phone (Formatted)']||'',type=l['Buyer/Seller']||'—',status=l['SMS Status']||'Pending',temp=l['Lead Temperature']||'',phase=l['Phase']||'';
    const sb={'Sent':'<span class="badge badge-sent">Sent</span>','Pending':'<span class="badge badge-pending">Pending</span>','Opted Out':'<span class="badge badge-optout">Opted Out</span>'}[status]||`<span class="badge badge-pending">${status}</span>`;
    const tb2={'Hot':'<span class="badge badge-hot">🔥</span>','Warm':'<span class="badge badge-warm">🌡️</span>','Cold':'<span class="badge badge-cold">❄️</span>'}[temp]||'';
    const pb2={'Phase 1':'<span class="badge badge-p1">P1</span>','Phase 2':'<span class="badge badge-p2">P2</span>','Phase 3':'<span class="badge badge-p3">P3</span>'}[phase]||'';
    return `<tr class="${currentPhone===phone?'selected':''}" onclick="selectLead('${phone}','${name.replace(/'/g,"\\'")}')"><td>${name}</td><td style="color:#94a3b8;font-size:12px;">${phone}</td><td style="font-size:12px;">${type}</td><td>${pb2}</td><td>${sb}</td><td>${tb2}</td></tr>`;
  }).join('');
}

async function selectLead(phone,name){
  currentPhone=phone;
  document.getElementById('conv-title').textContent=name;
  document.getElementById('reply-box').style.display='block';
  const d=await(await fetch(`/api/conversation/${encodeURIComponent(phone)}`)).json();
  const el=document.getElementById('conv-messages');
  let html='';
  if(d.initial_message&&d.initial_message!='nan')html+=`<div class="msg msg-out"><div class="msg-label">Sarah (Agent)</div><div class="msg-bubble">${d.initial_message}</div></div>`;
  if(d.reply_text&&d.reply_text!='nan')html+=`<div class="msg msg-in"><div class="msg-label">Lead</div><div class="msg-bubble">${d.reply_text}</div></div>`;
  if(d.history)d.history.forEach(m=>{html+=`<div class="msg ${m.role==='assistant'?'msg-out':'msg-in'}"><div class="msg-label">${m.role==='assistant'?'Sarah (Agent)':'Lead'}</div><div class="msg-bubble">${m.content}</div></div>`;});
  el.innerHTML=html||'<div style="padding:30px;text-align:center;color:#475569;font-size:13px;">No messages yet</div>';
  el.scrollTop=el.scrollHeight;
  loadLeads();
}

async function sendManualReply(){
  const msg=document.getElementById('reply-text').value.trim();
  if(!msg||!currentPhone)return;
  const d=await(await fetch('/api/reply',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({phone:currentPhone,message:msg})})).json();
  if(d.success){document.getElementById('reply-text').value='';showToast('Message sent!');selectLead(currentPhone,document.getElementById('conv-title').textContent);}
  else showToast('Failed to send',true);
}

function getBroadcastFilters(){
  return{
    phase:[...document.querySelectorAll('.bc-phase:checked')].map(e=>e.value),
    buyer_seller:[...document.querySelectorAll('.bc-type:checked')].map(e=>e.value),
    temperature:[...document.querySelectorAll('.bc-temp:checked')].map(e=>e.value),
    city:document.getElementById('bc-city').value.trim()
  };
}

async function previewBroadcast(){
  const d=await(await fetch('/api/broadcast/preview',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(getBroadcastFilters())})).json();
  if(d.error){showToast(d.error,true);return;}
  document.getElementById('broadcast-preview').innerHTML=`<strong>${d.count} leads</strong> will receive this message.<br><small style="color:#64748b;">Sample names: ${d.sample.join(', ')}</small>`;
}

async function sendBroadcast(){
  const msg=document.getElementById('broadcast-msg').value.trim();
  if(!msg){showToast('Please write a message first',true);return;}
  if(!confirm('Send broadcast to selected leads?'))return;
  const d=await(await fetch('/api/broadcast/send',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({message:msg,filters:getBroadcastFilters()})})).json();
  d.error?showToast(d.error,true):showToast('📢 '+d.message);
}

async function triggerCampaign(){
  if(!confirm('Launch campaign now?'))return;
  await fetch('/trigger');
  showToast('🚀 Campaign launched!');
  setTimeout(loadStats,3000);
}

async function togglePause(){
  const d=await(await fetch('/api/pause',{method:'POST'})).json();
  showToast(d.status==='paused'?'⏸ Campaign paused':'▶ Campaign resumed');
  loadStats();
}

function setFilter(f,el){currentFilter=f;document.querySelectorAll('.filter-btn').forEach(b=>b.classList.remove('active'));el.classList.add('active');loadLeads();}
function searchLeads(){clearTimeout(searchTimer);searchTimer=setTimeout(loadLeads,300);}

async function uploadLeads(input){
  const file=input.files[0];if(!file)return;
  const fd=new FormData();fd.append('file',file);
  showToast('Uploading...');
  const d=await(await fetch('/api/upload',{method:'POST',body:fd})).json();
  d.error?showToast(d.error,true):showToast('✅ '+d.message);
  loadStats();loadLeads();input.value='';
}

function showToast(msg,err=false){
  const t=document.getElementById('toast');
  t.textContent=msg;t.className='toast'+(err?' error':'');
  t.style.display='block';
  setTimeout(()=>t.style.display='none',4000);
}

document.getElementById('add-lead-modal').addEventListener('click',function(e){
  if(e.target===this)closeAddLead();
});

loadStats();loadLeads();setInterval(()=>{loadStats();loadLeads();},30000);
</script></body></html>"""

scheduler_thread = threading.Thread(target=scheduler_loop)
scheduler_thread.daemon = True
scheduler_thread.start()
print("[AGENT] Background scheduler started")
