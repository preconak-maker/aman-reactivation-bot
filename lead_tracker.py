import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import os
from config import LEADS_FILE, SHEET_NAME


def load_leads() -> pd.DataFrame:
    df = pd.read_excel(LEADS_FILE, sheet_name=SHEET_NAME)
    return df


def get_pending_leads(df: pd.DataFrame) -> pd.DataFrame:
    return df[df["SMS Status"] == "Pending"].copy()


def get_followup_leads(df: pd.DataFrame) -> pd.DataFrame:
    """Leads sent but no reply after FOLLOWUP_DAYS."""
    from config import FOLLOWUP_DAYS
    now = datetime.now()
    mask = (
        (df["SMS Status"] == "Sent") &
        (df["Reply Received"] == "No") &
        (df["SMS Sent At"] != "") &
        (df["SMS Sent At"].notna())
    )
    candidates = df[mask].copy()
    if candidates.empty:
        return candidates
    candidates["SMS Sent At"] = pd.to_datetime(candidates["SMS Sent At"], errors="coerce")
    overdue = candidates[(now - candidates["SMS Sent At"]).dt.days >= FOLLOWUP_DAYS]
    return overdue


def update_lead_sent(phone: str, message: str):
    """Mark lead as Sent in Excel."""
    _update_cell(phone, {
        "SMS Status": "Sent",
        "SMS Sent At": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "SMS Message Sent": message
    })


def update_lead_reply(phone: str, reply_text: str, temperature: str = ""):
    """Log incoming reply to Excel."""
    _update_cell(phone, {
        "Reply Received": "Yes",
        "Reply Text": reply_text,
        "Lead Temperature": temperature
    })


def update_lead_optout(phone: str):
    """Mark lead as opted out."""
    _update_cell(phone, {
        "SMS Status": "Opted Out",
        "Follow Up Required": "No"
    })


def _update_cell(phone: str, updates: dict):
    """Generic Excel row updater by phone number."""
    wb = load_workbook(LEADS_FILE)
    ws = wb[SHEET_NAME]

    # Build header map
    headers = {cell.value: cell.column for cell in ws[1]}
    phone_col = headers.get("Phone (Formatted)")

    for row in ws.iter_rows(min_row=2):
        cell_val = str(row[phone_col - 1].value or "").strip()
        if cell_val == str(phone).strip():
            for col_name, value in updates.items():
                col_idx = headers.get(col_name)
                if col_idx:
                    ws.cell(row=row[0].row, column=col_idx, value=value)
            break

    wb.save(LEADS_FILE)
